"""
Integre CyberAgressionAdo-Large (Zenodo 14770265) dans training/text/datasets/.

Chaque fichier xlsx contient plusieurs feuilles (re-annotations).
Schema : ID, TIME, NAME, TEXT, ROLE, HATE, TARGET, VERBAL_ABUSE, INTENTION, CONTEXT, SENTIMENT.

Mapping -> 5 classes Sentinel :
- HATE=OAG + ROLE in (bully, bully_support) + TARGET=victim  -> harassment (4)
- HATE=OAG + ROLE in (bully, bully_support) ailleurs          -> rage (2)
- HATE=CAG + ROLE in (bully, bully_support)                    -> harassment (4)
- HATE=OAG + ROLE in (victim, victim_support)                  -> anger (1)
- HATE=NAG + SENTIMENT in (NEU, POS)                           -> neutral (0)
- reste                                                         -> skip (ambigu)

Dedup global contre les fichiers existants (toxic/*.jsonl + neutral/*.txt) et
interne (meme texte vu plusieurs fois a cause des re-annotations).

Sorties :
    training/text/datasets/toxic/train_cyberagression_large.jsonl
    training/text/datasets/neutral/train_cyberagression_large.txt
"""
from __future__ import annotations

import glob
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
AI_ROOT = SCRIPT_DIR.parent
DATA_DIR = AI_ROOT / "training" / "text" / "datasets"
TOXIC_DIR = DATA_DIR / "toxic"
NEUTRAL_DIR = DATA_DIR / "neutral"
XLSX_DIR = SCRIPT_DIR / "cyberagression_large"

LABEL_NAMES = {0: "neutral", 1: "anger", 2: "rage", 3: "threat", 4: "harassment"}


def norm_key(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip().lower())


def load_existing_keys() -> set[str]:
    seen: set[str] = set()
    if TOXIC_DIR.exists():
        for p in TOXIC_DIR.glob("*.jsonl"):
            for line in p.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    seen.add(norm_key(json.loads(line)["text"]))
                except (json.JSONDecodeError, KeyError):
                    pass
    if NEUTRAL_DIR.exists():
        for p in NEUTRAL_DIR.glob("*.txt"):
            for line in p.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line:
                    seen.add(norm_key(line))
    return seen


def map_row(role: str, hate: str, target: str, sentiment: str) -> int | None:
    role = (role or "").strip().lower()
    hate = (hate or "").strip().upper()
    target = (target or "").strip().lower()
    sentiment = (sentiment or "").strip().upper()

    is_bully = role in {"bully", "bully_support"}
    is_victim = role in {"victim", "victim_support"}

    if hate == "OAG" and is_bully:
        return 4 if target == "victim" else 2
    if hate == "CAG" and is_bully:
        return 4
    if hate == "OAG" and is_victim:
        return 1
    if hate == "NAG" and sentiment in {"NEU", "POS"}:
        return 0
    return None


def main() -> None:
    if not XLSX_DIR.exists():
        raise SystemExit(f"Absent : {XLSX_DIR}")

    print("Index existant...")
    seen = load_existing_keys()
    print(f"  {len(seen)} textes indexes\n")

    toxic_out: list[dict] = []
    neutral_out: list[str] = []
    stats = Counter()
    dupes_ext = 0
    dupes_int = 0
    total_msgs = 0

    local_seen: set[str] = set()

    for fp in sorted(glob.glob(str(XLSX_DIR / "*.xlsx"))):
        wb = load_workbook(fp, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if not header:
                continue
            # indices : ID,TIME,NAME,TEXT,ROLE,HATE,TARGET,VERBAL_ABUSE,INTENTION,CONTEXT,SENTIMENT
            for r in it:
                if not r or len(r) < 11 or r[3] is None:
                    continue
                text = re.sub(r"\s+", " ", str(r[3]).strip())
                if not (3 <= len(text) <= 500):
                    continue
                total_msgs += 1
                key = norm_key(text)
                if key in seen:
                    dupes_ext += 1
                    continue
                if key in local_seen:
                    dupes_int += 1
                    continue
                label = map_row(r[4], r[5], r[6], r[10])
                if label is None:
                    continue
                local_seen.add(key)
                seen.add(key)
                if label == 0:
                    neutral_out.append(text)
                else:
                    toxic_out.append({"text": text, "label": label})
                stats[label] += 1

    print(f"Messages bruts lus    : {total_msgs}")
    print(f"Doublons externes     : {dupes_ext}")
    print(f"Doublons internes     : {dupes_int}")
    print(f"Toxic retenus         : {len(toxic_out)}")
    print(f"Neutral retenus       : {len(neutral_out)}")
    print("\nRepartition :")
    for lb in sorted(stats):
        print(f"  {LABEL_NAMES[lb]:12s}: {stats[lb]}")

    if toxic_out:
        out = TOXIC_DIR / "train_cyberagression_large.jsonl"
        out.parent.mkdir(parents=True, exist_ok=True)
        with out.open("w", encoding="utf-8") as f:
            for e in toxic_out:
                f.write(json.dumps(e, ensure_ascii=False) + "\n")
        print(f"\n  -> {out}")

    if neutral_out:
        out = NEUTRAL_DIR / "train_cyberagression_large.txt"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text("\n".join(neutral_out), encoding="utf-8")
        print(f"  -> {out}")


if __name__ == "__main__":
    main()
